from openpyxl import Workbook
wb= Workbook() # workbook / sheet
ws= wb.active # active sheet is
testdata=[['Name','Age'],["Anusha",33],["aksh",4],["ragh",38]]
for x in testdata:
    ws.append(x)
    wb.save("namenage.xlsx")

"""for i in range(1,6):
    for j in range(1,3):
        ws.cell(row=i,column=j).value=i+j"""


ws.cell(row=1, column=1).value=" Writing to specified cell"
ws.cell(row=1, column=2).value=" row1column2"
for i in range(2,6):
     for j in range(1,3):
         ws.cell(row=i,column=j).value=((i-1)*2)+(j)

# wb.save("demoexcel.xlsx")

