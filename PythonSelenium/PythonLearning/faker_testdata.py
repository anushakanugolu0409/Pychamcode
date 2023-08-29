from faker import Faker
from openpyxl import Workbook , load_workbook
from faker.providers import DynamicProvider
import pandas as pd

"""wb=Workbook()
ws=wb.active
fake=Faker()
medical_professions_provider = DynamicProvider(
     provider_name="medical_profession",
     elements=["dr.", "doctor", "nurse", "surgeon", "clerk"],
)
fake.add_provider(medical_professions_provider)

ws.cell(row=1,column=1).value="NAME"
ws.cell(row=1,column=2).value="EMAIL"
ws.cell(row=1,column=3).value="ADDRESS"
ws.cell(row=1,column=4).value="medical_profession"
for i in range (2,100):
    for j in range (1,5):
        ws.cell(row=i,column=1).value= Faker().name()
        ws.cell(row=i,column=2).value= Faker().email()
        ws.cell(row=i,column=3).value= Faker().address()
        ws.cell(row=i, column=4).value= fake.medical_profession()
wb.save("testdatagenarate.xlsx")"""

wb= load_workbook(filename="testdatagenarate.xlsx")
ws=wb['Sheet']
a= []
rowcount= ws.max_row
for i in range (1, rowcount+1):
    resu= ws.cell(row=i, column=1).value
    a.append(resu)
"""print(a)
print(len(set(a))== len(a))"""
df= pd.DataFrame(a) # convert list to data frame

colname = df.set_axis(['Names1'],axis=1)
p= df.value_counts()>1 # check if any duplicates are there
#  print(sum(p)) check total count of duplicate values
for x in range (len(df)):
    if (df.value_counts()>1).any() == False: #any() is used to return the boolen value
        print(df)





