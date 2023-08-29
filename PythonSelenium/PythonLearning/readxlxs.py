from openpyxl import Workbook,load_workbook
wb=load_workbook(filename="demoexcel.xlsx")
ws=wb['Sheet']
print(ws.cell(2,2).value)
rowcount =ws.max_row
colcount = ws.max_column

for i in range(1,rowcount+1):
    for j in range(1,colcount+1):
        print(ws.cell(row=i,column=j).value,end='')
        print("\n")


